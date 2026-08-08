"""Excel içe/dışa aktarım ve normalizasyon testleri."""

import datetime as dt
import io

import openpyxl
import pytest

from app.excel import sema
from app.excel.ice_aktar import _para, _tarih, oku


# --------------------------------------------------------------------------- #
# Normalizasyon
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("ham,beklenen", [
    ("Ip Kamera", "IP Kamera"),
    ("İp Kamera", "IP Kamera"),
    ("ip Kamera", "IP Kamera"),
    ("İp Kamera (değiştirildi) - iade edildi", "IP Kamera"),
    ("Dizüstü Bilgisayar", "Dizüstü Bilgisayar"),
    ("Monitör", "Monitör"),
    ("Yansıtıcı", "Projeksiyon"),
    ("HARDİSK", "Harddisk"),
    ("Firewall", "Güvenlik Duvarı"),
    ("", "Diğer"),
])
def test_cihaz_tipi_normalizasyonu(ham, beklenen):
    """Türkçe I/İ farkı yüzünden aynı tür ayrı kategorilere düşmemeli."""
    assert sema.cihaz_tipi_normalle(ham) == beklenen


@pytest.mark.parametrize("ad,kisi", [
    ("Ahmet Yılmaz", True),
    ("ALİ ARSLAN", True),                     # büyük harf tek başına sinyal değil
    ("Ömer Faruk Koçak", True),
    ("Eser Pehlivan (Site Plus)", True),      # parantezli ek bilgi
    ("Sunucu Odası", False),
    ("İNŞAAT SAHASI", False),
    ("yenisi ile değiştirildi", False),
    ("STAFF KISMI İŞÇİ KOĞUŞU GÜVENLİK", False),
    ("Boşta", False),
    ("?", False),
])
def test_kisi_mi(ad, kisi):
    """'Kullanıcı Adı' sütununda yer/not yazanlar kişi olarak açılmamalı."""
    assert sema.kisi_mi(ad) is kisi


def test_sadelestir_turkce():
    assert sema._sadelestir("ŞANTİYE") == sema._sadelestir("şantiye")
    assert sema._sadelestir("Işık") == sema._sadelestir("IŞIK")


# --------------------------------------------------------------------------- #
# Değer dönüştürücüler
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("giris,beklenen", [
    ("1.250,50", 1250.50),   # Türkçe biçim
    ("1250.50", 1250.50),
    (28500, 28500.0),
    ("0", None),             # Excel'de boş yerine 0 yazılmış
    ("", None),
    (None, None),
])
def test_para_ayristirma(giris, beklenen):
    assert _para(giris) == beklenen


@pytest.mark.parametrize("giris,beklenen", [
    ("2017-04-21", dt.date(2017, 4, 21)),
    ("21.04.2017", dt.date(2017, 4, 21)),
    (dt.datetime(2017, 4, 21, 10, 30), dt.date(2017, 4, 21)),
    ("saçma", None),
    (0, None),
])
def test_tarih_ayristirma(giris, beklenen):
    assert _tarih(giris) == beklenen


# --------------------------------------------------------------------------- #
# Dosya okuma
# --------------------------------------------------------------------------- #
def _excel_uret(satirlar, basliklar=None, baslik_yaz=True) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    basliklar = basliklar or sema.STANDART_SUTUNLAR
    if baslik_yaz:
        ws.append(basliklar)
    for s in satirlar:
        ws.append([s.get(b) for b in basliklar])
    tampon = io.BytesIO()
    wb.save(tampon)
    return tampon.getvalue()


def test_basliklı_dosya_okunur():
    veri = _excel_uret([
        {"Cihaz Tipi": "Dizüstü Bilgisayar", "Cihaz NO": "N100",
         "Serial": "SN1", "Marka": "ASUS", "Kullanıcı Adı": "Ahmet Yılmaz",
         "İşlemci (Bütün)": "Intel i5", "Ram Kapasitesi (mB)": "16 Gb"},
    ])
    sonuc = oku(veri)
    assert sonuc["toplam"] == 1
    s = sonuc["satirlar"][0]
    assert s["asset_tag"] == "N100"
    assert s["cihaz_tipi"] == "Dizüstü Bilgisayar"
    assert s["kisi_mi"] is True
    assert s["ozellikler"]["İşlemci"]["İşlemci (Bütün)"] == "Intel i5"
    assert s["ozellikler"]["Bellek"]["Ram Kapasitesi (mB)"] == "16 Gb"


def test_basliksiz_dosya_standart_siraya_gore_okunur():
    """Bazı dosyalarda başlık satırı yok; sütun düzeni standarttır."""
    veri = _excel_uret(
        [{"Cihaz Tipi": "Monitör", "Cihaz NO": "M028", "Serial": "ALMA9"}],
        baslik_yaz=False,
    )
    sonuc = oku(veri)
    assert sonuc["toplam"] == 1
    assert sonuc["satirlar"][0]["asset_tag"] == "M028"
    assert any("başlık satırı yok" in u for u in sonuc["uyarilar"])


def test_bos_satirlar_atlanir():
    veri = _excel_uret([
        {"Cihaz NO": "A1", "Serial": "S1"},
        {},                       # tamamen boş
        {"Marka": "Dell"},        # etiket ve seri yok -> atlanmalı
    ])
    assert oku(veri)["toplam"] == 1


def test_tekrarli_etiket_uyarisi():
    veri = _excel_uret([
        {"Cihaz NO": "N1", "Serial": "S1"},
        {"Cihaz NO": "N1", "Serial": "S2"},
    ])
    sonuc = oku(veri)
    assert any("birden fazla" in u for u in sonuc["uyarilar"])


def test_bilinmeyen_sutun_ek_bilgide_saklanir():
    basliklar = sema.STANDART_SUTUNLAR + ["Garanti Firması"]
    veri = _excel_uret(
        [{"Cihaz NO": "X1", "Serial": "S9", "Garanti Firması": "ABC Servis"}],
        basliklar=basliklar,
    )
    s = oku(veri)["satirlar"][0]
    assert s["ozellikler"]["Ek Bilgi"]["Garanti Firması"] == "ABC Servis"


# --------------------------------------------------------------------------- #
# Uçlar
# --------------------------------------------------------------------------- #
CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_ice_aktarim_ucu(client):
    veri = _excel_uret([
        {"Cihaz Tipi": "Dizüstü Bilgisayar", "Cihaz NO": "N500", "Serial": "SN500",
         "Marka": "Dell", "Model": "Latitude", "Kullanıcı Adı": "Ayşe Demir",
         "Bulunduğu Yer": "ŞANTİYE", "Fiyat (TL)": "28.500,00",
         "Fatura Tarihi": "2024-03-15", "İşlemci (Bütün)": "Intel i7"},
    ])
    on = client.post("/excel/oku", files={"file": ("t.xlsx", io.BytesIO(veri), CT)})
    assert on.status_code == 200, on.text
    assert on.json()["toplam"] == 1

    r = client.post("/excel/aktar", json={"satirlar": on.json()["satirlar"]})
    assert r.status_code == 200, r.text
    assert r.json()["eklenen"] == 1
    assert r.json()["atlanan"] == 0

    varlik = next(a for a in client.get("/assets").json() if a["asset_tag"] == "N500")
    assert varlik["purchase_cost"] == 28500.0
    assert varlik["purchase_date"] == "2024-03-15"
    assert varlik["assigned_type"] == "user"
    assert varlik["custom"]["İşlemci"]["İşlemci (Bütün)"] == "Intel i7"


def test_json_gidip_gelen_tarih_metni_kabul_edilir(client):
    """Önizleme JSON'dan geçtiği için tarih/sayı metne dönüşür; aktarım
    bunları geri çevirmeli (aksi hâlde 'SQLite Date type' hatası)."""
    r = client.post("/excel/aktar", json={"satirlar": [{
        "asset_tag": "JSON-1", "serial": "JS1", "cihaz_tipi": "Monitör",
        "purchase_date": "2024-01-05",     # metin!
        "purchase_cost": "1.500,75",       # metin!
        "ozellikler": {}, "kisi_mi": False,
    }]})
    assert r.status_code == 200
    assert r.json()["atlanan"] == 0, r.json()["hatalar"]
    a = next(x for x in client.get("/assets").json() if x["asset_tag"] == "JSON-1")
    assert a["purchase_date"] == "2024-01-05"
    assert a["purchase_cost"] == 1500.75


def test_ayni_seri_ile_guncellenir(client):
    veri = _excel_uret([{"Cihaz NO": "U1", "Serial": "AYNI", "Marka": "HP"}])
    on = client.post("/excel/oku", files={"file": ("t.xlsx", io.BytesIO(veri), CT)})
    client.post("/excel/aktar", json={"satirlar": on.json()["satirlar"]})

    veri2 = _excel_uret([{"Cihaz NO": "U1", "Serial": "AYNI", "Marka": "Dell"}])
    on2 = client.post("/excel/oku", files={"file": ("t.xlsx", io.BytesIO(veri2), CT)})
    r = client.post("/excel/aktar", json={"satirlar": on2.json()["satirlar"]})
    assert r.json()["guncellenen"] == 1 and r.json()["eklenen"] == 0


def test_disa_aktarim(client):
    client.post("/assets", json={"asset_tag": "EXP-1", "name": "Test", "serial": "E1"})
    r = client.get("/excel/disa-aktar.xlsx")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    basliklar = [c.value for c in ws[1]]
    assert basliklar == sema.STANDART_SUTUNLAR, "Dışa aktarım aynı sütun düzeninde olmalı"
    assert any(ws.cell(i, basliklar.index("Cihaz NO") + 1).value == "EXP-1"
               for i in range(2, ws.max_row + 1))


def test_sablon_indirilir(client):
    r = client.get("/excel/sablon.xlsx")
    assert r.status_code == 200
    ws = openpyxl.load_workbook(io.BytesIO(r.content)).active
    assert [c.value for c in ws[1]] == sema.STANDART_SUTUNLAR


def test_gecersiz_dosya_reddedilir(client):
    r = client.post("/excel/oku",
                    files={"file": ("a.txt", io.BytesIO(b"metin"), "text/plain")})
    assert r.status_code == 400


def test_ice_aktarim_editor_ister(viewer_client):
    veri = _excel_uret([{"Cihaz NO": "V1", "Serial": "VS1"}])
    r = viewer_client.post("/excel/oku",
                           files={"file": ("t.xlsx", io.BytesIO(veri), CT)})
    assert r.status_code == 403


# --------------------------------------------------------------------------- #
# Eksik marka: mevcut model kaydı yeniden aktarımda tamamlanmalı
# --------------------------------------------------------------------------- #
def test_markasiz_model_yeniden_aktarimda_markalanir(client):
    """Snipe-IT'ten markasız gelen model, Excel'de marka varsa dolmalı.

    Marka cihazda değil modelde durur; model adla bulunduğu için eskiden
    mevcut kayıt olduğu gibi kullanılıyor ve marka sonsuza dek boş kalıyordu.
    """
    # Üretimdeki durum: model var, markası yok
    kat = client.post("/categories", json={"name": "Dizüstü"}).json()
    mdl = client.post("/models", json={"name": "N439",
                                       "category_id": kat["id"]}).json()
    assert mdl["manufacturer_id"] is None

    veri = _excel_uret([{"Cihaz Tipi": "Dizüstü", "Cihaz NO": "FRM-0002",
                         "Serial": "SN-0002", "Marka": "HP", "Model": "N439"}])
    on = client.post("/excel/oku", files={"file": ("t.xlsx", io.BytesIO(veri), CT)})
    r = client.post("/excel/aktar", json={"satirlar": on.json()["satirlar"]})
    assert r.json()["atlanan"] == 0, r.json()["hatalar"]

    # Yeni model açılmamalı, mevcut kayıt markalanmalı
    modeller = [m for m in client.get("/models").json() if m["name"] == "N439"]
    assert len(modeller) == 1
    marka = client.get(f"/manufacturers/{modeller[0]['manufacturer_id']}").json()
    assert marka["name"] == "HP"

    # Cihaz detayında da görünmeli
    varlik = next(a for a in client.get("/assets").json()
                  if a["asset_tag"] == "FRM-0002")
    assert client.get(f"/detay/asset/{varlik['id']}").json()["kunye"]["marka"] == "HP"


def test_dolu_marka_baska_bir_aktarimla_ezilmez(client):
    """Aynı model adı başka markada da kullanılabilir; dolu değer korunur."""
    hp = client.post("/manufacturers", json={"name": "HP"}).json()
    client.post("/models", json={"name": "X1", "manufacturer_id": hp["id"]})

    veri = _excel_uret([{"Cihaz NO": "K-1", "Serial": "K1", "Marka": "Dell",
                         "Model": "X1"}])
    on = client.post("/excel/oku", files={"file": ("t.xlsx", io.BytesIO(veri), CT)})
    client.post("/excel/aktar", json={"satirlar": on.json()["satirlar"]})

    mdl = next(m for m in client.get("/models").json() if m["name"] == "X1")
    assert client.get(
        f"/manufacturers/{mdl['manufacturer_id']}").json()["name"] == "HP"

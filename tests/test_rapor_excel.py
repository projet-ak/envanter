"""Biçimli Excel raporları: içerik, stil ve uç davranışı."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.rapor import KOYU


@pytest.fixture
def sahne(client):
    lok = client.post("/locations", json={"name": "ŞANTİYE U070",
                                          "proje_kodu": "U070"}).json()
    kisi = client.post("/users", json={"first_name": "Rapor",
                                       "last_name": "Testi",
                                       "employee_num": "R-1",
                                       "department": "BT"}).json()
    cihaz = client.post("/assets", json={
        "asset_tag": "RPR-1", "name": "Rapor Laptopu", "serial": "SN-RPR",
        "location_id": lok["id"], "purchase_cost": 25000,
        "purchase_date": "2026-01-15"}).json()
    client.post(f"/assets/{cihaz['id']}/checkout",
                json={"assigned_type": "user", "assigned_id": kisi["id"]})
    client.post("/assets", json={"asset_tag": "RPR-2", "name": "Boşta Monitör",
                                 "location_id": lok["id"]})
    client.post("/accessories", json={"name": "Rapor Klavyesi", "qty": 2,
                                      "min_qty": 5})
    client.post("/ag/urunler", json={
        "tur": "switch", "asset_tag": "RPR-SW", "marka": "HUAWEI",
        "location_id": lok["id"], "ozellikler": {"Port Sayısı": "24"}})
    return {"lok": lok, "kisi": kisi}


def _kitap(client, tip):
    r = client.get("/reports/excel", params={"tip": tip})
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]
    assert ".xlsx" in r.headers["content-disposition"]
    return load_workbook(BytesIO(r.content))


def test_genel_rapor_tum_sayfalari_icerir(client, sahne):
    wb = _kitap(client, "genel")
    assert wb.sheetnames == ["Özet", "Cihazlar", "Zimmetler", "Lokasyonlar",
                             "Stok", "Sistem Ürünleri"]


def test_cihazlar_sayfasi_bicimli(client, sahne, monkeypatch, tmp_path):
    # Logo dosyası ortamda olabilir de olmayabilir de; bu test logosuz düzeni
    # sınar — yol bilerek var olmayan bir dosyaya çevrilir.
    from app import rapor
    monkeypatch.setattr(rapor, "LOGO_YOLU", tmp_path / "yok.png")

    ws = _kitap(client, "cihazlar")["Cihazlar"]
    # 1. satır kurum başlığı, 2. satır tablo başlığı
    assert "Cihaz Listesi" in ws["A1"].value
    assert ws["A2"].value == "Cihaz No"
    assert ws["A2"].fill.start_color.rgb.endswith(KOYU)   # kurumsal yeşil
    assert ws["A2"].font.bold and ws["A2"].font.color.rgb == "00FFFFFF"
    assert ws.freeze_panes == "A3"
    assert ws.auto_filter.ref and ws.auto_filter.ref.startswith("A2:")

    satirlar = {r[0]: r for r in ws.iter_rows(min_row=3, values_only=True)}
    assert "RPR-1" in satirlar
    kayit = satirlar["RPR-1"]
    assert kayit[2] == "Rapor Laptopu"
    assert kayit[10] == "Rapor Testi"          # zimmetli kişi
    # Tarih gerçek tarih hücresi, para TL biçiminde
    basliklar = [c.value for c in ws[2]]
    tarih_sutun = basliklar.index("Alım Tarihi") + 1
    para_sutun = basliklar.index("Alım Bedeli") + 1
    hucre_satiri = [r for r in ws.iter_rows(min_row=3) if r[0].value == "RPR-1"][0]
    assert hucre_satiri[tarih_sutun - 1].number_format == "DD.MM.YYYY"
    assert "₺" in hucre_satiri[para_sutun - 1].number_format
    assert hucre_satiri[para_sutun - 1].value == 25000.0


def test_zimmet_raporu_yalnizca_zimmetlileri_verir(client, sahne):
    ws = _kitap(client, "zimmet")["Zimmetler"]
    etiketler = [r[0] for r in ws.iter_rows(min_row=3, values_only=True)]
    assert "RPR-1" in etiketler
    assert "RPR-2" not in etiketler            # boştaki cihaz girmez
    kayit = [r for r in ws.iter_rows(min_row=3, values_only=True)
             if r[0] == "RPR-1"][0]
    assert kayit[11] == "R-1" and kayit[12] == "BT"


def test_lokasyon_raporu_sayilari_dogru(client, sahne):
    ws = _kitap(client, "lokasyon")["Lokasyonlar"]
    kayit = [r for r in ws.iter_rows(min_row=3, values_only=True)
             if r[0] == "ŞANTİYE U070"][0]
    # RPR-1 + RPR-2 + RPR-SW = 3 cihaz, 1 zimmetli, 2 boşta
    assert kayit[1] == "U070"
    assert (kayit[2], kayit[3], kayit[4]) == (3, 1, 2)


def test_stok_raporu_dusuk_stogu_isaretler(client, sahne):
    ws = _kitap(client, "stok")["Stok"]
    kayit = [r for r in ws.iter_rows(min_row=3, values_only=True)
             if r[1] == "Rapor Klavyesi"][0]
    assert kayit[0] == "Aksesuar" and kayit[3] == 2 and kayit[5] == "DÜŞÜK"


def test_sistem_raporu_ozellikleri_yazar(client, sahne):
    ws = _kitap(client, "sistem")["Sistem Ürünleri"]
    kayit = [r for r in ws.iter_rows(min_row=3, values_only=True)
             if r[2] == "RPR-SW"][0]
    assert kayit[0] == "Ağ Ürünleri" and kayit[1] == "Switch"
    assert "Port Sayısı: 24" in kayit[10]


def test_bilinmeyen_tip_reddedilir(client):
    assert client.get("/reports/excel", params={"tip": "yok"}).status_code == 400


def test_rapor_giris_ister(anon_client):
    assert anon_client.get("/reports/excel").status_code == 401


def test_son_islemler(client, sahne):
    r = client.get("/reports/son-islemler").json()
    assert r, "işlem geçmişi boş olmamalı"
    assert {"hedef", "eylem", "tarih"} <= set(r[0])
    eylemler = {x["eylem"] for x in r}
    assert "zimmetlendi" in eylemler or "eklendi" in eylemler


def test_logo_varsa_rapora_gomulur(client, sahne, monkeypatch, tmp_path):
    """logo-rapor.png konduğunda başlık sağa kayar ve görsel gömülür."""
    from PIL import Image as PILImage

    from app import rapor

    yol = tmp_path / "logo-rapor.png"
    PILImage.new("RGBA", (200, 140), (0, 61, 53, 255)).save(yol)
    monkeypatch.setattr(rapor, "LOGO_YOLU", yol)

    ws = _kitap(client, "cihazlar")["Cihazlar"]
    assert ws["C1"].value and "Cihaz Listesi" in ws["C1"].value
    assert len(ws._images) == 1, "logo çalışma kitabına gömülmemiş"


def test_logo_dar_sayfada_atlanir_genel_rapor_bozulmaz(client, sahne,
                                                       monkeypatch, tmp_path):
    """Özet sayfası 2 sütun: logo C1 birleştirmesi orada geçersiz aralık
    üretiyordu (C1:B1 → 500). Dar sayfada logo atlanmalı, kalanlarda durmalı."""
    from PIL import Image as PILImage

    from app import rapor

    yol = tmp_path / "logo-rapor.png"
    PILImage.new("RGBA", (200, 140), (0, 61, 53, 255)).save(yol)
    monkeypatch.setattr(rapor, "LOGO_YOLU", yol)

    wb = _kitap(client, "genel")            # 500 dönerse _kitap assert'ü patlar
    assert "Genel Envanter Raporu" in wb["Özet"]["A1"].value   # dar: logosuz
    assert len(wb["Özet"]._images) == 0
    assert "Cihaz Listesi" in wb["Cihazlar"]["C1"].value       # geniş: logolu
    assert len(wb["Cihazlar"]._images) == 1


def test_secili_kimliklerle_rapor_daraltilir(client, sahne):
    """Listeden seçim: ids verilince yalnızca o cihazlar rapora girer."""
    hepsi = client.get("/assets").json()
    secilen = next(a for a in hepsi if a["asset_tag"] == "RPR-1")
    ws = _kitap(client, f"cihazlar")["Cihazlar"] if False else None
    r = client.get("/reports/excel",
                   params={"tip": "cihazlar", "ids": str(secilen["id"])})
    assert r.status_code == 200
    ws = load_workbook(BytesIO(r.content))["Cihazlar"]
    etiketler = [x[0] for x in ws.iter_rows(min_row=3, values_only=True)]
    assert etiketler == ["RPR-1"]

    assert client.get("/reports/excel",
                      params={"tip": "cihazlar", "ids": "a,b"}).status_code == 400

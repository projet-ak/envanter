"""Biçimli Excel raporları.

CSV dışa aktarımından farkı: bunlar **sunuma hazır** dosyalar — kurum başlığı,
renkli tablo başlıkları (ERN koyu yeşili), dondurulmuş üst satır, süzgeç okları,
zebra satırlar, otomatik sütun genişliği, Türkçe tarih/para biçimleri.

Rapor türleri (`olustur(db, tip)`):

    cihazlar   tüm varlıklar, tam künye
    zimmet     yalnızca zimmetli cihazlar + kişi bilgileri
    lokasyon   şantiye/lokasyon özeti (cihaz, zimmetli, boşta)
    stok       aksesuar / sarf / bileşen / lisans, adetlerle
    sistem     sistem ürünleri (ağ, yangın, alarm, geçiş, kantar) + özellikler
    genel      hepsi tek çalışma kitabında, başına Özet sayfası
"""

from __future__ import annotations

import datetime as dt
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app import ag, models
from app.config import settings

KOYU = "003D35"          # ERN kurumsal koyu yeşil
ZEBRA = "F2F7F5"         # açık zebra satır
CIZGI = "C9D6D1"

_BASLIK_FONT = Font(bold=True, color="FFFFFF", size=11)
_BASLIK_DOLGU = PatternFill("solid", start_color=KOYU)
_ZEBRA_DOLGU = PatternFill("solid", start_color=ZEBRA)
_INCE = Side(style="thin", color=CIZGI)
_KENAR = Border(left=_INCE, right=_INCE, top=_INCE, bottom=_INCE)

# Rapor başlığına gömülecek kurum logosu (yeşil versiyon, beyaz zemine göre).
# Dosya yoksa raporlar logosuz ama aynı düzende üretilir.
LOGO_YOLU = Path(__file__).parent / "static" / "logo-rapor.png"

TARIH_BICIMI = "DD.MM.YYYY"
PARA_BICIMI = '#,##0.00 "₺"'

RAPOR_ADLARI = {
    "genel": "Genel Envanter Raporu",
    "cihazlar": "Cihaz Listesi",
    "zimmet": "Zimmet Raporu",
    "lokasyon": "Lokasyon Raporu",
    "stok": "Stok Raporu",
    "sistem": "Sistem Ürünleri Raporu",
}


def _sayfa(wb: Workbook, ad: str, baslik: str, sutunlar: list[str],
           satirlar: list[list], *, para_sutunlar: set[int] = frozenset(),
           tarih_sutunlar: set[int] = frozenset()) -> None:
    """Tek biçimli sayfa: kurum başlığı, stilli tablo, süzgeç, donmuş satır."""
    ws = wb.create_sheet(ad)
    ws.sheet_properties.tabColor = KOYU

    son_sutun = get_column_letter(len(sutunlar))
    # Logo A1-B1 alanına oturur; başlık C1'den başlar. 3 sütundan dar
    # sayfada (örn. Özet) C1'den birleştirme geçersiz aralık üretir
    # (C1:B1) — orada logo atlanır, başlık A1'de kalır.
    logolu = LOGO_YOLU.exists() and len(sutunlar) >= 4
    ilk = "C" if logolu else "A"
    ws.merge_cells(f"{ilk}1:{son_sutun}1")
    ust = ws[f"{ilk}1"]
    ust.value = (f"{settings.org_name} — {baslik} — "
                 f"{dt.date.today():%d.%m.%Y}")
    ust.font = Font(bold=True, size=13, color=KOYU)
    ust.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34 if logolu else 26
    if logolu:
        img = XLImage(str(LOGO_YOLU))
        oran = 40 / img.height             # ~40px yükseklik, oran korunur
        img.height, img.width = 40, int(img.width * oran)
        ws.add_image(img, "A1")

    for i, etiket in enumerate(sutunlar, start=1):
        h = ws.cell(row=2, column=i, value=etiket)
        h.font = _BASLIK_FONT
        h.fill = _BASLIK_DOLGU
        h.border = _KENAR
        h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    for r, satir in enumerate(satirlar, start=3):
        for c, deger in enumerate(satir, start=1):
            hucre = ws.cell(row=r, column=c, value=deger)
            hucre.border = _KENAR
            if r % 2 == 1:                       # 3, 5, 7… zebra
                hucre.fill = _ZEBRA_DOLGU
            if c in para_sutunlar and deger is not None:
                hucre.number_format = PARA_BICIMI
            if c in tarih_sutunlar and deger is not None:
                hucre.number_format = TARIH_BICIMI

    # Sütun genişliği veriden: kısa kalmasın, ekranı da yutmasın
    for i, etiket in enumerate(sutunlar, start=1):
        en = len(str(etiket))
        for satir in satirlar[:400]:
            v = satir[i - 1]
            if v is not None:
                en = max(en, len(str(v)))
        ws.column_dimensions[get_column_letter(i)].width = min(max(en + 3, 11), 46)

    ws.freeze_panes = "A3"
    if satirlar:
        ws.auto_filter.ref = f"A2:{son_sutun}{len(satirlar) + 2}"


# --------------------------------------------------------------------------- #
# Veri toplayıcılar
# --------------------------------------------------------------------------- #
def _haritalar(db: Session) -> dict:
    return {
        "model": {m.id: m for m in db.scalars(select(models.AssetModel)).all()},
        "kategori": {k.id: k.name for k in db.scalars(select(models.Category)).all()},
        "marka": {m.id: m.name
                  for m in db.scalars(select(models.Manufacturer)).all()},
        "lokasyon": {l.id: l for l in db.scalars(select(models.Location)).all()},
        "durum": {s.id: s.name for s in db.scalars(select(models.StatusLabel)).all()},
        "kisi": {k.id: k for k in db.scalars(select(models.User)).all()},
    }


def _cihaz_satirlari(db: Session, h: dict, *, sadece_zimmetli: bool) -> list[list]:
    satirlar = []
    for a in db.scalars(select(models.Asset).order_by(models.Asset.asset_tag)).all():
        kisi = h["kisi"].get(a.assigned_user_id) if a.assigned_user_id else None
        if sadece_zimmetli and kisi is None:
            continue
        mdl = h["model"].get(a.model_id)
        lok = h["lokasyon"].get(a.location_id)
        satir = [
            a.asset_tag, a.demirbas_no, a.name,
            h["kategori"].get(mdl.category_id) if mdl else None,
            h["marka"].get(mdl.manufacturer_id) if mdl else None,
            mdl.name if mdl else None, a.serial,
            lok.name if lok else None,
            lok.proje_kodu if lok else None,
            h["durum"].get(a.status_id),
            kisi.full_name if kisi else None,
        ]
        if sadece_zimmetli:
            satir += [kisi.employee_num, kisi.department, kisi.job_title,
                      a.last_checkout.date() if a.last_checkout else None]
        else:
            satir += [a.ip_address, a.purchase_date,
                      float(a.purchase_cost) if a.purchase_cost else None,
                      a.warranty_end]
        satirlar.append(satir)
    return satirlar


def _cihazlar_sayfasi(wb: Workbook, db: Session, h: dict) -> None:
    _sayfa(wb, "Cihazlar", RAPOR_ADLARI["cihazlar"],
           ["Cihaz No", "Demirbaş No", "Ad", "Tür", "Marka", "Model", "Seri No",
            "Lokasyon", "Proje", "Durum", "Zimmetli Kişi", "IP",
            "Alım Tarihi", "Alım Bedeli", "Garanti Bitiş"],
           _cihaz_satirlari(db, h, sadece_zimmetli=False),
           para_sutunlar={14}, tarih_sutunlar={13, 15})


def _zimmet_sayfasi(wb: Workbook, db: Session, h: dict) -> None:
    _sayfa(wb, "Zimmetler", RAPOR_ADLARI["zimmet"],
           ["Cihaz No", "Demirbaş No", "Ad", "Tür", "Marka", "Model", "Seri No",
            "Lokasyon", "Proje", "Durum", "Zimmetli Kişi", "Sicil No",
            "Departman", "Unvan", "Zimmet Tarihi"],
           _cihaz_satirlari(db, h, sadece_zimmetli=True),
           tarih_sutunlar={15})


def _lokasyon_sayfasi(wb: Workbook, db: Session, h: dict) -> None:
    sayilar: dict[int | None, list[int]] = {}
    for a in db.scalars(select(models.Asset)).all():
        kayit = sayilar.setdefault(a.location_id, [0, 0])
        kayit[0] += 1
        if a.assigned_user_id or a.assigned_type is not None:
            kayit[1] += 1
    satirlar = []
    for lok_id, (toplam, zimmetli) in sorted(
            sayilar.items(), key=lambda x: -x[1][0]):
        lok = h["lokasyon"].get(lok_id)
        satirlar.append([
            lok.name if lok else "(belirtilmemiş)",
            lok.proje_kodu if lok else None,
            toplam, zimmetli, toplam - zimmetli,
        ])
    _sayfa(wb, "Lokasyonlar", RAPOR_ADLARI["lokasyon"],
           ["Lokasyon", "Proje Kodu", "Cihaz", "Zimmetli", "Boşta"], satirlar)


def _stok_sayfasi(wb: Workbook, db: Session) -> None:
    satirlar = []
    for tablo, tur in ((models.Accessory, "Aksesuar"),
                       (models.Consumable, "Sarf Malzeme"),
                       (models.Component, "Bileşen")):
        for s in db.scalars(select(tablo).order_by(tablo.name)).all():
            dusuk = (s.min_qty is not None and s.qty is not None
                     and s.qty <= s.min_qty)
            satirlar.append([tur, s.name, getattr(s, "model_number", None),
                             s.qty, s.min_qty, "DÜŞÜK" if dusuk else None])
    for l in db.scalars(select(models.License).order_by(models.License.name)).all():
        satirlar.append(["Lisans", l.name, l.license_key, l.seats, None, None])
    _sayfa(wb, "Stok", RAPOR_ADLARI["stok"],
           ["Tür", "Ad", "Model No / Anahtar", "Adet", "Asgari", "Uyarı"],
           satirlar)


def _sistem_sayfasi(wb: Workbook, db: Session) -> None:
    satirlar = []
    for u in ag.urunler(db):
        bilgi = ag.TURLER.get(u["tur"], {})
        ozellikler = " · ".join(f"{k}: {v}" for k, v in u["ozellikler"].items())
        satirlar.append([
            ag.AILELER.get(bilgi.get("aile"), {}).get("ad"),
            bilgi.get("ad"), u["asset_tag"], u["marka"], u["model"],
            u["serial"], u["lokasyon"], u["proje_kodu"], u["durum"],
            u["zimmetli"], ozellikler or None,
        ])
    _sayfa(wb, "Sistem Ürünleri", RAPOR_ADLARI["sistem"],
           ["Aile", "Tür", "Cihaz No", "Marka", "Model", "Seri No", "Lokasyon",
            "Proje", "Durum", "Zimmetli", "Teknik Özellikler"], satirlar)


def _ozet_sayfasi(wb: Workbook, db: Session, h: dict) -> None:
    varliklar = db.scalars(select(models.Asset)).all()
    zimmetli = sum(1 for a in varliklar if a.assigned_type is not None)
    durum_sayilari: dict[str, int] = {}
    kategori_sayilari: dict[str, int] = {}
    for a in varliklar:
        durum_sayilari[h["durum"].get(a.status_id) or "(belirtilmemiş)"] = \
            durum_sayilari.get(h["durum"].get(a.status_id)
                               or "(belirtilmemiş)", 0) + 1
        mdl = h["model"].get(a.model_id)
        kat = h["kategori"].get(mdl.category_id) if mdl else None
        kategori_sayilari[kat or "(belirtilmemiş)"] = \
            kategori_sayilari.get(kat or "(belirtilmemiş)", 0) + 1

    satirlar = [
        ["Toplam cihaz", len(varliklar)],
        ["Zimmetli", zimmetli],
        ["Boşta", len(varliklar) - zimmetli],
        ["Personel", len(h["kisi"])],
        ["Lokasyon", len(h["lokasyon"])],
        ["", None],
        ["— Cihaz Tipine Göre —", None],
        *[[k, n] for k, n in sorted(kategori_sayilari.items(),
                                    key=lambda x: -x[1])],
        ["", None],
        ["— Duruma Göre —", None],
        *[[k, n] for k, n in sorted(durum_sayilari.items(),
                                    key=lambda x: -x[1])],
    ]
    _sayfa(wb, "Özet", RAPOR_ADLARI["genel"], ["Gösterge", "Değer"], satirlar)


def olustur(db: Session, tip: str) -> bytes:
    """Raporu üretir ve xlsx baytlarını döner. Bilinmeyen tip -> KeyError."""
    RAPOR_ADLARI[tip]                       # bilinmeyen tipte erken patla
    wb = Workbook()
    wb.remove(wb.active)                    # boş varsayılan sayfa kalmasın
    h = _haritalar(db)

    if tip in ("genel",):
        _ozet_sayfasi(wb, db, h)
    if tip in ("genel", "cihazlar"):
        _cihazlar_sayfasi(wb, db, h)
    if tip in ("genel", "zimmet"):
        _zimmet_sayfasi(wb, db, h)
    if tip in ("genel", "lokasyon"):
        _lokasyon_sayfasi(wb, db, h)
    if tip in ("genel", "stok"):
        _stok_sayfasi(wb, db)
    if tip in ("genel", "sistem"):
        _sistem_sayfasi(wb, db)

    tampon = BytesIO()
    wb.save(tampon)
    return tampon.getvalue()

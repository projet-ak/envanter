"""Excel içe/dışa aktarım uçları (kurumun 65 sütunlu envanter formatı)."""

from __future__ import annotations

import io

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response
from sqlalchemy import select
from sqlalchemy.orm import Session

from app import models
from app.auth import get_current_user, require_editor
from app.database import get_db
from app.excel import sema
from app.excel.ice_aktar import aktar as _aktar
from app.excel.ice_aktar import oku as _oku

router = APIRouter(prefix="/excel", tags=["Excel"])

MAX_BOYUT = 25 * 1024 * 1024  # 25 MB


@router.post("/oku", dependencies=[Depends(require_editor)])
async def excel_oku(file: UploadFile = File(...)):
    """Excel dosyasını ayrıştırıp önizleme döndürür (hiçbir şey kaydetmez)."""
    ad = (file.filename or "").lower()
    if not ad.endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Yalnızca .xlsx / .xlsm dosyaları desteklenir")

    veri = await file.read()
    if not veri:
        raise HTTPException(400, "Dosya boş")
    if len(veri) > MAX_BOYUT:
        raise HTTPException(413, "Dosya 25 MB'tan büyük olamaz")

    try:
        sonuc = _oku(veri)
    except Exception as exc:
        raise HTTPException(400, f"Dosya okunamadı: {type(exc).__name__}: {exc}")

    if not sonuc["toplam"]:
        raise HTTPException(
            400, "Dosyada içe aktarılabilir satır bulunamadı. "
                 "'Cihaz NO' veya 'Serial' sütunu dolu olmalı."
        )

    # Önizlemede ilk 50 satırı göster; tamamı aktarımda kullanılır
    return {
        "toplam": sonuc["toplam"],
        "tipler": sonuc["tipler"],
        "kisi_sayisi": sonuc["kisi_sayisi"],
        "uyarilar": sonuc["uyarilar"],
        "ornekler": sonuc["satirlar"][:50],
        "satirlar": sonuc["satirlar"],
    }


@router.post("/aktar", dependencies=[Depends(require_editor)])
def excel_aktar(payload: dict, db: Session = Depends(get_db)):
    """Önizlemeden onaylanan satırları envantere işler."""
    satirlar = payload.get("satirlar") or []
    if not satirlar:
        raise HTTPException(400, "Aktarılacak satır yok")
    guncelle = bool(payload.get("guncelle", True))
    return _aktar(db, satirlar, guncelle=guncelle)


@router.get("/disa-aktar.xlsx", dependencies=[Depends(get_current_user)])
def excel_disa_aktar(db: Session = Depends(get_db)):
    """Tüm varlıkları kurumun Excel formatında dışa aktarır."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Envanter"

    basliklar = sema.STANDART_SUTUNLAR
    ws.append(basliklar)
    baslik_stili = Font(bold=True, color="FFFFFF")
    dolgu = PatternFill("solid", fgColor="2F5597")
    for h in ws[1]:
        h.font = baslik_stili
        h.fill = dolgu
        h.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    # İlişkileri tek seferde çöz (N+1 sorgu olmasın)
    def harita(model):
        return {n.id: n.name for n in db.scalars(select(model)).all()}

    lokasyonlar = harita(models.Location)
    tedarikciler = harita(models.Supplier)
    sirketler = harita(models.Company)
    modeller = {m.id: m for m in db.scalars(select(models.AssetModel)).all()}
    ureticiler = harita(models.Manufacturer)
    kategoriler = harita(models.Category)
    kisiler = {
        k.id: k for k in db.scalars(select(models.User)).all()
    }

    for a in db.scalars(select(models.Asset).order_by(models.Asset.asset_tag)):
        mdl = modeller.get(a.model_id)
        kisi = kisiler.get(a.assigned_user_id) if a.assigned_user_id else None
        ozel = a.custom or {}

        def oz(grup: str, alan: str):
            return (ozel.get(grup) or {}).get(alan)

        satir_veri = {
            "Cihaz Tipi": kategoriler.get(mdl.category_id) if mdl else None,
            "Cihaz NO": a.asset_tag,
            "Serial": a.serial,
            "IFS KOD": a.muhasebe_kodu,
            "Bulunduğu Yer": lokasyonlar.get(a.location_id),
            "Kullanıcı Adı": " ".join(filter(None, [kisi.first_name, kisi.last_name]))
                             if kisi else None,
            "Kullanılan Birim": kisi.department if kisi else None,
            "Unvan": kisi.job_title if kisi else None,
            "IP": a.ip_address,
            "Marka": ureticiler.get(mdl.manufacturer_id) if mdl else None,
            "Model": mdl.name if mdl else None,
            "Fatura Tarihi": a.purchase_date,
            "Fatura No": a.fatura_no,
            "Tedarikçi Firma Adı": tedarikciler.get(a.supplier_id),
            "Alınan Şirket": sirketler.get(a.company_id),
            "Fiyat (TL)": float(a.purchase_cost) if a.purchase_cost else None,
            "Açıklama": a.notes,
        }
        # Teknik özellikleri geri yaz
        for grup, alanlar in sema.OZELLIK_GRUPLARI.items():
            for alan in alanlar:
                if satir_veri.get(alan) is None:
                    satir_veri[alan] = oz(grup, alan)

        ws.append([satir_veri.get(b) for b in basliklar])

    # Sütun genişlikleri
    for i, b in enumerate(basliklar, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = min(max(12, len(b) + 2), 40)

    tampon = io.BytesIO()
    wb.save(tampon)
    return Response(
        content=tampon.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="envanter.xlsx"'},
    )


@router.get("/sablon.xlsx", dependencies=[Depends(get_current_user)])
def excel_sablon():
    """Boş şablon dosyası (doğru sütun başlıklarıyla)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Envanter"
    ws.append(sema.STANDART_SUTUNLAR)
    for h in ws[1]:
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="2F5597")
        h.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for i, b in enumerate(sema.STANDART_SUTUNLAR, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = min(max(12, len(b) + 2), 40)

    tampon = io.BytesIO()
    wb.save(tampon)
    return Response(
        content=tampon.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="envanter-sablon.xlsx"'},
    )

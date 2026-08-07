"""Excel envanter dosyasını okuyup varlıklara dönüştürür.

Akış: `oku()` dosyayı ayrıştırıp önizleme üretir (hiçbir şey yazılmaz),
`aktar()` onaylanan satırları veritabanına işler.
"""

from __future__ import annotations

import datetime as dt
import io
import re
from typing import Any

import openpyxl
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app import models
from app.excel import sema


# --------------------------------------------------------------------------- #
# Değer dönüştürücüler
# --------------------------------------------------------------------------- #
def _metin(deger: Any) -> str | None:
    if deger is None:
        return None
    s = str(deger).strip()
    # Excel'de boş yerine "0" veya "-" yazılmış olabiliyor
    if s in ("", "0", "-", "yok", "Yok", "YOK", "None", "nan"):
        return None
    return re.sub(r"\s+", " ", s)


def _tarih(deger: Any) -> dt.date | None:
    if deger is None or deger == 0:
        return None
    if isinstance(deger, dt.datetime):
        return deger.date()
    if isinstance(deger, dt.date):
        return deger
    s = str(deger).strip()
    for kalip in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s[:10], kalip).date()
        except ValueError:
            continue
    return None


def _para(deger: Any) -> float | None:
    if deger in (None, "", 0, "0"):
        return None
    if isinstance(deger, (int, float)):
        return float(deger) or None
    s = str(deger).strip().replace("₺", "").replace("$", "").replace(" ", "")
    # Türkçe biçim: 1.250,50 -> 1250.50
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s) or None
    except ValueError:
        return None


# --------------------------------------------------------------------------- #
# Okuma / önizleme
# --------------------------------------------------------------------------- #
def _baslik_satiri(ws) -> int:
    """Başlık satırını bulur (ilk 5 satırda en çok dolu hücreye sahip olan)."""
    en_iyi, en_iyi_i = 0, 1
    for i in range(1, min(6, ws.max_row + 1)):
        dolu = sum(1 for c in ws[i] if c.value not in (None, ""))
        if dolu > en_iyi:
            en_iyi, en_iyi_i = dolu, i
    return en_iyi_i


def oku(icerik: bytes) -> dict:
    """Excel dosyasını ayrıştırıp satırları ve özeti döndürür (yazmaz)."""
    wb = openpyxl.load_workbook(io.BytesIO(icerik), data_only=True)
    satirlar: list[dict] = []
    uyarilar: list[str] = []
    bilinmeyen_sutunlar: set[str] = set()

    for ws in wb.worksheets:
        bas_i = _baslik_satiri(ws)
        basliklar = [
            str(c.value).strip() if c.value is not None else ""
            for c in ws[bas_i]
        ]

        if sema.CIHAZ_TIPI in basliklar or "Cihaz NO" in basliklar:
            veri_baslangic = bas_i + 1
        else:
            # Başlık satırı yok: bazı dosyalarda veri 1. satırdan başlar ama
            # sütun düzeni standarttır. Standart sıraya göre yorumla.
            basliklar = sema.STANDART_SUTUNLAR
            veri_baslangic = 1
            uyarilar.append(
                f"'{ws.title}': başlık satırı yok, standart sütun sırası varsayıldı"
            )

        idx = {b: i for i, b in enumerate(basliklar) if b}
        bilinmeyen_sutunlar |= {b for b in idx if b not in sema.BILINEN}

        for satir_no, r in enumerate(
            ws.iter_rows(min_row=veri_baslangic, values_only=True),
            start=veri_baslangic
        ):
            if not any(v not in (None, "") for v in r):
                continue

            def al(basl: str) -> Any:
                i = idx.get(basl)
                return r[i] if i is not None and i < len(r) else None

            etiket = _metin(al("Cihaz NO"))
            seri = _metin(al("Serial"))
            if not etiket and not seri:
                continue  # tamamen boş kayıt

            kayit: dict = {
                "sayfa": ws.title,
                "satir": satir_no,
                "asset_tag": etiket,
                "serial": seri,
                "cihaz_tipi": sema.cihaz_tipi_normalle(_metin(al(sema.CIHAZ_TIPI)) or ""),
                "marka": _metin(al(sema.MARKA)),
                "model": _metin(al(sema.MODEL)),
                "kullanici": sema.ad_normalle(_metin(al(sema.KULLANICI)) or ""),
                "lokasyon": _metin(al(sema.LOKASYON)),
                "birim": _metin(al(sema.BIRIM)),
                "unvan": _metin(al(sema.UNVAN)),
                "sirket": _metin(al(sema.SIRKET)),
                "tedarikci": _metin(al(sema.TEDARIKCI)),
                "purchase_date": _tarih(al(sema.FATURA_TARIHI)),
                "purchase_cost": _para(al(sema.FIYAT_TL)),
            }
            for excel_basi, alan in sema.ALAN_ESLEME.items():
                if alan not in kayit:
                    kayit[alan] = _metin(al(excel_basi))

            # Teknik özellikler -> custom
            ozellikler: dict[str, dict[str, str]] = {}
            for grup, alanlar in sema.OZELLIK_GRUPLARI.items():
                grup_veri = {a: _metin(al(a)) for a in alanlar}
                grup_veri = {k: v for k, v in grup_veri.items() if v}
                if grup_veri:
                    ozellikler[grup] = grup_veri
            # Bilinmeyen sütunlar da kaybolmasın
            ek = {b: _metin(r[i]) for b, i in idx.items()
                  if b not in sema.BILINEN and i < len(r)}
            ek = {k: v for k, v in ek.items() if v}
            if ek:
                ozellikler["Ek Bilgi"] = ek
            kayit["ozellikler"] = ozellikler
            kayit["kisi_mi"] = sema.kisi_mi(kayit["kullanici"])

            satirlar.append(kayit)

    # Özet
    tipler: dict[str, int] = {}
    kisiler: set[str] = set()
    for s in satirlar:
        tipler[s["cihaz_tipi"]] = tipler.get(s["cihaz_tipi"], 0) + 1
        if s["kisi_mi"]:
            kisiler.add(s["kullanici"])

    etiketsiz = sum(1 for s in satirlar if not s["asset_tag"])
    if etiketsiz:
        uyarilar.append(
            f"{etiketsiz} satırda Cihaz NO yok — seri no veya sıra ile etiket üretilecek"
        )

    # Dosya içi tekrar eden etiketler
    sayac: dict[str, int] = {}
    for s in satirlar:
        if s["asset_tag"]:
            sayac[s["asset_tag"]] = sayac.get(s["asset_tag"], 0) + 1
    tekrar = {k: v for k, v in sayac.items() if v > 1}
    if tekrar:
        uyarilar.append(
            f"{len(tekrar)} Cihaz NO birden fazla satırda geçiyor "
            f"(örn. {', '.join(list(tekrar)[:3])}) — sonuna -2, -3 eklenecek"
        )
    if bilinmeyen_sutunlar:
        uyarilar.append(
            f"Tanınmayan {len(bilinmeyen_sutunlar)} sütun 'Ek Bilgi' altında saklanacak"
        )

    return {
        "toplam": len(satirlar),
        "tipler": dict(sorted(tipler.items(), key=lambda x: -x[1])),
        "kisi_sayisi": len(kisiler),
        "uyarilar": uyarilar,
        "satirlar": satirlar,
    }


# --------------------------------------------------------------------------- #
# Aktarım
# --------------------------------------------------------------------------- #
class _Onbellek:
    """Ada göre kayıt arayan/oluşturan önbellek.

    Neden veritabanı sorgusu değil:
    SQLite'ın LOWER() fonksiyonu yalnızca ASCII harfleri çevirir
    ("ŞANTİYE" -> "Şantİye"), Python'unki ise tam Unicode uygular. Bu
    uyuşmazlık yüzünden `func.lower(...) == ad.lower()` karşılaştırması
    Türkçe adlarda ASLA eşleşmiyor ve her satır yeni kayıt oluşturuyordu
    (486 "ŞANTİYE" lokasyonu). Karşılaştırmayı Python tarafında yaparak
    veritabanı harf-çevirme davranışına bağımlılığı kaldırıyoruz.
    """

    def __init__(self, db: Session):
        self.db = db
        self._tablolar: dict[type, dict[str, Any]] = {}

    def _yukle(self, model) -> dict[str, Any]:
        if model not in self._tablolar:
            self._tablolar[model] = {
                sema._sadelestir(n.name): n
                for n in self.db.scalars(select(model)).all()
                if n.name
            }
        return self._tablolar[model]

    def al(self, model, ad: str | None, **ekstra):
        ad = (ad or "").strip()
        if not ad:
            return None
        harita = self._yukle(model)
        anahtar = sema._sadelestir(ad)
        if anahtar in harita:
            return harita[anahtar]
        nesne = model(name=ad, **ekstra)
        self.db.add(nesne)
        self.db.flush()
        harita[anahtar] = nesne
        return nesne

    def kullanici(self, tam_ad: str) -> models.User | None:
        tam_ad = sema.ad_normalle(tam_ad)
        if not tam_ad:
            return None
        if models.User not in self._tablolar:
            self._tablolar[models.User] = {
                sema._sadelestir(
                    " ".join(filter(None, [k.first_name, k.last_name]))
                ): k
                for k in self.db.scalars(select(models.User)).all()
            }
        harita = self._tablolar[models.User]
        anahtar = sema._sadelestir(tam_ad)
        if anahtar in harita:
            return harita[anahtar]

        parcalar = tam_ad.split()
        kisi = models.User(
            first_name=parcalar[0],
            last_name=" ".join(parcalar[1:]) or None,
        )
        self.db.add(kisi)
        self.db.flush()
        harita[anahtar] = kisi
        return kisi


def _benzersiz_etiket(db: Session, istenen: str | None, seri: str | None,
                      kullanilan: set[str], sira: int) -> str:
    taban = (istenen or seri or f"EXL-{sira:05d}").strip()
    aday = taban
    n = 1
    while aday in kullanilan or db.scalar(
        select(models.Asset.id).where(models.Asset.asset_tag == aday)
    ):
        n += 1
        aday = f"{taban}-{n}"
    kullanilan.add(aday)
    return aday


def aktar(db: Session, satirlar: list[dict], *, varsayilan_durum_id: int | None = None,
          guncelle: bool = True) -> dict:
    """Önizlemeden gelen satırları veritabanına işler.

    `guncelle=True` ise seri numarası eşleşen mevcut varlık güncellenir,
    aksi hâlde her satır yeni kayıt olur.
    """
    eklenen = guncellenen = atlanan = 0
    hatalar: list[str] = []
    kullanilan_etiketler: set[str] = set()
    onbellek = _Onbellek(db)

    if varsayilan_durum_id is None:
        durum = db.scalar(
            select(models.StatusLabel).where(
                models.StatusLabel.type == models.StatusType.deployable
            )
        )
        varsayilan_durum_id = durum.id if durum else None

    for sira, s in enumerate(satirlar, start=1):
        try:
            kategori = onbellek.al(models.Category, s.get("cihaz_tipi"))
            uretici = onbellek.al(models.Manufacturer, s.get("marka"))
            tedarikci = onbellek.al(models.Supplier, s.get("tedarikci"))
            sirket = onbellek.al(models.Company, s.get("sirket"))

            # Lokasyon: "Bulunduğu Yer" yoksa ve kullanıcı bir yer adıysa onu kullan
            lokasyon_adi = s.get("lokasyon")
            if not lokasyon_adi and not s.get("kisi_mi") and s.get("kullanici"):
                lokasyon_adi = s["kullanici"]
            lokasyon = onbellek.al(models.Location, lokasyon_adi)
            # Excel'deki "Kullanılan Birim" (U023, U026…) proje kodudur;
            # lokasyonda boşsa doldur.
            if lokasyon is not None and s.get("birim") and not lokasyon.proje_kodu:
                lokasyon.proje_kodu = s["birim"]

            model_adi = s.get("model") or (
                f"{s.get('marka')} {s.get('cihaz_tipi')}".strip()
                if s.get("marka") else s.get("cihaz_tipi")
            )
            model = onbellek.al(
                models.AssetModel, model_adi,
                category_id=kategori.id if kategori else None,
                manufacturer_id=uretici.id if uretici else None,
            ) if model_adi else None

            # Mevcut kayıt var mı? (seri no en güvenilir anahtar)
            varlik = None
            if guncelle and s.get("serial"):
                varlik = db.scalar(
                    select(models.Asset).where(models.Asset.serial == s["serial"])
                )

            yeni_mi = varlik is None
            if yeni_mi:
                varlik = models.Asset(
                    asset_tag=_benzersiz_etiket(
                        db, s.get("asset_tag"), s.get("serial"),
                        kullanilan_etiketler, sira)
                )
                db.add(varlik)

            varlik.name = " ".join(filter(None, [s.get("marka"), s.get("model")])) \
                or s.get("cihaz_tipi")
            varlik.serial = s.get("serial") or varlik.serial
            varlik.model_id = model.id if model else varlik.model_id
            varlik.location_id = lokasyon.id if lokasyon else varlik.location_id
            varlik.supplier_id = tedarikci.id if tedarikci else varlik.supplier_id
            varlik.company_id = sirket.id if sirket else varlik.company_id
            varlik.status_id = varlik.status_id or varsayilan_durum_id
            for alan in ("muhasebe_kodu", "fatura_no", "ip_address", "notes"):
                if s.get(alan):
                    setattr(varlik, alan, s[alan])
            # Önizleme JSON üzerinden geldiğinde tarih/sayı metne dönüşür;
            # veritabanına yazmadan önce tipleri geri çevir.
            if s.get("purchase_date"):
                tarih = _tarih(s["purchase_date"])
                if tarih:
                    varlik.purchase_date = tarih
            if s.get("purchase_cost"):
                bedel = _para(s["purchase_cost"])
                if bedel:
                    varlik.purchase_cost = bedel
            if s.get("warranty_end"):
                garanti = _tarih(s["warranty_end"])
                if garanti:
                    varlik.warranty_end = garanti

            # Teknik özellikler
            mevcut = dict(varlik.custom or {})
            mevcut.update(s.get("ozellikler") or {})
            varlik.custom = mevcut

            db.flush()

            # Zimmet
            if s.get("kisi_mi") and s.get("kullanici"):
                kisi = onbellek.kullanici(s["kullanici"])
                if kisi:
                    if s.get("birim") and not kisi.department:
                        kisi.department = s["birim"]
                    if s.get("unvan") and not kisi.job_title:
                        kisi.job_title = s["unvan"]
                    varlik.assigned_type = models.AssignedType.user
                    varlik.assigned_user_id = kisi.id
                    varlik.assigned_location_id = None

            db.add(models.ActivityLog(
                action=models.ActivityAction.create if yeni_mi
                else models.ActivityAction.update,
                item_type="asset", item_id=varlik.id,
                note=f"Excel içe aktarım ({s.get('sayfa')} / satır {s.get('satir')})",
            ))

            if yeni_mi:
                eklenen += 1
            else:
                guncellenen += 1

        except Exception as exc:  # tek satır tüm aktarımı bozmasın
            db.rollback()
            atlanan += 1
            if len(hatalar) < 30:
                hatalar.append(
                    f"{s.get('sayfa')} satır {s.get('satir')}: {type(exc).__name__}: {exc}"
                )

    db.commit()
    return {
        "eklenen": eklenen,
        "guncellenen": guncellenen,
        "atlanan": atlanan,
        "hatalar": hatalar,
    }

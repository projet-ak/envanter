#!/usr/bin/env python3
"""Snipe-IT dışa aktarımını mevcut veriyle karşılaştırır ve eksikleri doldurur.

Snipe-IT'in "Export" düğmesi `.xls` uzantılı ama aslında HTML tablosu olan bir
dosya üretir; bu betik hem onu hem gerçek `.xlsx` dosyasını okur.

Ne yapar:
  1. Dosyadaki her satırı **Demirbaş Etiketi** ile (o tutmazsa seri no ile)
     sistemdeki cihazla eşleştirir.
  2. Yalnızca sistemde **BOŞ** olan alanları doldurur. Dolu bir değeri asla
     ezmez — sistem artık kaynak, dosya eski bir fotoğraf.
  3. İki tarafta da dolu ama farklı olan alanları "çakışma" olarak listeler;
     bunlara dokunmaz, karar sizin.
  4. Lokasyon / durum / zimmet **hiç yazılmaz**: bunlar sistemde günlük
     değişen bilgiler, eski dosyadan yazmak yeni değişiklikleri geri alır.
     Yalnızca farklıysa rapora düşer.

Marka özel durumu: marka cihazda değil MODELDE durur. Snipe-IT'te de boşsa
`--marka-tahmini` aynı alım partisindeki (etiket önekindeki) kardeş cihazların
markasını kullanır — örn. FRM-0002-34543-… partisindeki cihazlar HP ise aynı
partideki markasız cihaz da HP sayılır. Varsayılan olarak en az **2** hemfikir
kardeş aranır (gerçek veride 34 denemede %100 isabet); 1'e düşürmek kapsamı
artırır ama isabet %96'ya iner.

Kullanım:
    ./.venv/bin/python scripts/snipeit-karsilastir.py disaaktarim.xls
    ./.venv/bin/python scripts/snipeit-karsilastir.py disaaktarim.xls --uygula
    ./.venv/bin/python scripts/snipeit-karsilastir.py disaaktarim.xls \
        --marka-tahmini --uygula
"""

from __future__ import annotations

import argparse
import html
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from sqlalchemy import select  # noqa: E402

from app import models  # noqa: E402
from app.database import SessionLocal  # noqa: E402
from app.excel.sema import _sadelestir  # noqa: E402
from app.ortam_uyari import uyar  # noqa: E402

M, Y, S, K, N = '\033[36m', '\033[32m', '\033[33m', '\033[31m', '\033[0m'

# Dosya sütunu -> ne olduğu
ETIKET = "Demirbaş Etiketi"
SERI = "Seri No"

# Cihazın kendi sütunlarına yazılabilecek alanlar: (dosya sütunu, alan)
VARLIK_ALANLARI = [
    ("Demirbaş Adı", "name"),
    (SERI, "serial"),
    ("Notlar", "notes"),
    ("IFS Cihaz Kodu", "muhasebe_kodu"),
]
# Modele yazılabilecekler
MODEL_ALANLARI = [("Model No.", "model_number")]
# Yalnızca raporlanan, asla yazılmayan alanlar
BILGI_ALANLARI = [("Konum", "lokasyon"), ("Durum", "durum"),
                  ("Çıkış Yapılmış Olan Kişi", "zimmetli")]


# --------------------------------------------------------------------------- #
# Dosya okuma
# --------------------------------------------------------------------------- #
def _html_tablo(metin: str) -> list[dict]:
    satirlar = re.findall(r"<tr[^>]*>(.*?)</tr>", metin, re.S | re.I)
    if not satirlar:
        return []

    def hucreler(tr: str) -> list[str]:
        return [html.unescape(re.sub(r"<[^>]+>", "", c)).strip()
                for c in re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", tr, re.S | re.I)]

    basliklar = hucreler(satirlar[0])
    return [dict(zip(basliklar, hucreler(t))) for t in satirlar[1:]]


def oku(yol: Path) -> list[dict]:
    """Snipe-IT dışa aktarımını okur (HTML görünümlü .xls ya da gerçek .xlsx)."""
    ham = yol.read_bytes()
    if ham[:2] == b"PK":                       # gerçek xlsx
        from openpyxl import load_workbook
        wb = load_workbook(yol, read_only=True, data_only=True)
        sayfa = wb[wb.sheetnames[0]]
        satirlar = list(sayfa.iter_rows(values_only=True))
        wb.close()
        if not satirlar:
            return []
        basliklar = [str(b or "").strip() for b in satirlar[0]]
        return [{b: ("" if h is None else str(h).strip())
                 for b, h in zip(basliklar, s)} for s in satirlar[1:]]
    return _html_tablo(ham.decode("utf-8", errors="replace"))


def _parti(etiket: str) -> str | None:
    """FRM-0002-34543-2552202208 -> FRM-0002-34543 (aynı alım partisi)."""
    p = (etiket or "").split("-")
    return "-".join(p[:-1]) if len(p) >= 3 else None


# --------------------------------------------------------------------------- #
# Karşılaştırma
# --------------------------------------------------------------------------- #
def _ad(nesne) -> str | None:
    return nesne.name if nesne is not None else None


def _bul(db, model, ad: str | None):
    """Ada göre kayıt (Türkçe duyarlı); yoksa None."""
    ad = (ad or "").strip()
    if not ad:
        return None
    aranan = _sadelestir(ad)
    for n in db.scalars(select(model)).all():
        if n.name and _sadelestir(n.name) == aranan:
            return n
    return None


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("dosya", type=Path, help="Snipe-IT dışa aktarım dosyası")
    ap.add_argument("--uygula", action="store_true",
                    help="boş alanları doldur (varsayılan: yalnızca rapor)")
    ap.add_argument("--marka-tahmini", action="store_true",
                    help="dosyada da boş olan markayı aynı partiden tahmin et")
    ap.add_argument("--en-az", type=int, default=2, metavar="N",
                    help="tahmin için gereken hemfikir kardeş sayısı (varsayılan 2)")
    args = ap.parse_args()

    if not args.dosya.exists():
        print(f"{K}Dosya bulunamadı: {args.dosya}{N}")
        return 1

    satirlar = [s for s in oku(args.dosya) if (s.get(ETIKET) or "").strip()]
    if not satirlar:
        print(f"{K}Dosyada satır okunamadı (beklenen sütun: '{ETIKET}'){N}")
        return 1

    uyar()
    db = SessionLocal()
    try:
        varliklar = db.scalars(select(models.Asset)).all()
        etikete_gore = {_sadelestir(a.asset_tag): a for a in varliklar if a.asset_tag}
        seriye_gore = {_sadelestir(a.serial): a for a in varliklar if a.serial}

        eslesen, dosyada_fazla = [], []
        for s in satirlar:
            a = (etikete_gore.get(_sadelestir(s[ETIKET]))
                 or seriye_gore.get(_sadelestir(s.get(SERI, ""))))
            (eslesen if a else dosyada_fazla).append((s, a))

        gorulen = {id(a) for _, a in eslesen if a}
        sistemde_fazla = [a for a in varliklar if id(a) not in gorulen]

        print(f"\n{M}Eşleştirme{N}")
        print(f"  Dosyadaki satır      : {len(satirlar)}")
        print(f"  Sistemdeki cihaz     : {len(varliklar)}")
        print(f"  {Y}Eşleşen              : {len(eslesen)}{N}")
        print(f"  Dosyada olup sistemde yok : {len(dosyada_fazla)}")
        print(f"  Sistemde olup dosyada yok : {len(sistemde_fazla)}")
        for s, _ in dosyada_fazla[:10]:
            print(f"      + {s[ETIKET]}  ({s.get('Model')})")
        if len(dosyada_fazla) > 10:
            print(f"      … ve {len(dosyada_fazla) - 10} satır daha")
        for a in sistemde_fazla[:10]:
            print(f"      - {a.asset_tag}")
        if len(sistemde_fazla) > 10:
            print(f"      … ve {len(sistemde_fazla) - 10} cihaz daha")

        # ---- Marka tahmini için parti haritası ---------------------------- #
        parti_markalar: dict[str, list[str]] = {}
        for s in satirlar:
            u = (s.get("Üretici") or "").strip()
            p = _parti(s[ETIKET])
            if u and p:
                parti_markalar.setdefault(p, []).append(u)

        def parti_tahmini(etiket: str) -> tuple[str, int] | None:
            kardesler = parti_markalar.get(_parti(etiket) or "", [])
            if len(set(kardesler)) == 1 and len(kardesler) >= args.en_az:
                return kardesler[0], len(kardesler)
            return None

        # ---- Alan alan karşılaştırma -------------------------------------- #
        doldurulacak: list[tuple] = []      # (varlık, "nerede", alan, yeni, kaynak)
        cakisma: list[tuple] = []
        bilgi_farki: list[tuple] = []

        for s, a in eslesen:
            mdl = db.get(models.AssetModel, a.model_id) if a.model_id else None

            for sutun, alan in VARLIK_ALANLARI:
                yeni = (s.get(sutun) or "").strip()
                if not yeni:
                    continue
                mevcut = getattr(a, alan, None)
                if not mevcut:
                    doldurulacak.append((a, "cihaz", alan, yeni, "dosya"))
                elif _sadelestir(str(mevcut)) != _sadelestir(yeni):
                    cakisma.append((a, alan, mevcut, yeni))

            if mdl is not None:
                for sutun, alan in MODEL_ALANLARI:
                    yeni = (s.get(sutun) or "").strip()
                    mevcut = getattr(mdl, alan, None)
                    if yeni and not mevcut:
                        doldurulacak.append((mdl, "model", alan, yeni, "dosya"))

                # Marka: modelde boşsa doldur
                if mdl.manufacturer_id is None:
                    uretici = (s.get("Üretici") or "").strip()
                    kaynak = "dosya"
                    if not uretici and args.marka_tahmini:
                        t = parti_tahmini(s[ETIKET])
                        if t:
                            uretici, kaynak = t[0], f"parti ({t[1]} kardeş)"
                    if uretici:
                        doldurulacak.append((mdl, "model", "marka", uretici, kaynak))

                # Kategori: modelde boşsa doldur
                if mdl.category_id is None and (s.get("Kategori") or "").strip():
                    doldurulacak.append((mdl, "model", "kategori",
                                         s["Kategori"].strip(), "dosya"))

            for sutun, ad in BILGI_ALANLARI:
                yeni = (s.get(sutun) or "").strip()
                if not yeni:
                    continue
                if ad == "lokasyon":
                    mevcut = _ad(db.get(models.Location, a.location_id)
                                 if a.location_id else None)
                elif ad == "durum":
                    mevcut = _ad(db.get(models.StatusLabel, a.status_id)
                                 if a.status_id else None)
                else:
                    k = (db.get(models.User, a.assigned_user_id)
                         if a.assigned_user_id else None)
                    mevcut = " ".join(filter(None, [k.first_name, k.last_name])) \
                        if k else None
                if mevcut and _sadelestir(mevcut) != _sadelestir(yeni):
                    bilgi_farki.append((a, ad, mevcut, yeni))

        # ---- Rapor -------------------------------------------------------- #
        marka_dolan = [d for d in doldurulacak if d[2] == "marka"]
        print(f"\n{M}Doldurulabilecek boş alanlar: {len(doldurulacak)}{N}")
        sayac: dict[str, int] = {}
        for _, _, alan, _, _ in doldurulacak:
            sayac[alan] = sayac.get(alan, 0) + 1
        for alan, n in sorted(sayac.items(), key=lambda x: -x[1]):
            print(f"    {alan:<16} {n:>4}")

        if marka_dolan:
            print(f"\n{M}Marka atanacak modeller (ilk 25){N}")
            for nesne, _, _, yeni, kaynak in marka_dolan[:25]:
                isaret = Y if kaynak == "dosya" else S
                print(f"  {isaret}→{N} {nesne.name:<24} ⇒ {yeni:<12} [{kaynak}]")
            if len(marka_dolan) > 25:
                print(f"    … ve {len(marka_dolan) - 25} model daha")

        if cakisma:
            print(f"\n{S}Çakışma — iki tarafta da dolu ve farklı "
                  f"({len(cakisma)}) — dokunulmadı{N}")
            for a, alan, mevcut, yeni in cakisma[:15]:
                print(f"  ? {a.asset_tag:<28} {alan}: "
                      f"sistem '{str(mevcut)[:28]}' ≠ dosya '{yeni[:28]}'")
            if len(cakisma) > 15:
                print(f"    … ve {len(cakisma) - 15} fark daha")

        if bilgi_farki:
            print(f"\n{M}Lokasyon / durum / zimmet farkları "
                  f"({len(bilgi_farki)}) — bilgi amaçlı, yazılmaz{N}")
            for a, ad, mevcut, yeni in bilgi_farki[:15]:
                print(f"  · {a.asset_tag:<28} {ad}: "
                      f"sistem '{str(mevcut)[:24]}' ≠ dosya '{yeni[:24]}'")
            if len(bilgi_farki) > 15:
                print(f"    … ve {len(bilgi_farki) - 15} fark daha")

        if not args.uygula:
            print(f"\n{M}Rapor modundasınız — hiçbir şey değişmedi.{N}")
            print("  Yazmak için komuta --uygula ekleyin.")
            if not args.marka_tahmini:
                print("  Dosyada da boş olan markaları partiden tahmin etmek "
                      "için: --marka-tahmini")
            return 0

        # ---- Yazma -------------------------------------------------------- #
        if not doldurulacak:
            print(f"\n{Y}✓ Doldurulacak boş alan yok — veri zaten güncel.{N}")
            return 0
        print(f"\n{S}Yazılıyor… (önce yedek aldığınızdan emin olun: "
              f"Ayarlar → Yedekleme){N}")
        for nesne, nerede, alan, yeni, _ in doldurulacak:
            if nerede == "cihaz":
                setattr(nesne, alan, yeni)
            elif alan == "marka":
                uretici = _bul(db, models.Manufacturer, yeni)
                if uretici is None:
                    uretici = models.Manufacturer(name=yeni)
                    db.add(uretici)
                    db.flush()
                nesne.manufacturer_id = uretici.id
            elif alan == "kategori":
                kat = _bul(db, models.Category, yeni)
                if kat is None:
                    kat = models.Category(name=yeni)
                    db.add(kat)
                    db.flush()
                nesne.category_id = kat.id
            else:
                setattr(nesne, alan, yeni)
        db.commit()
        print(f"\n{Y}✓ {len(doldurulacak)} alan dolduruldu "
              f"({len(marka_dolan)} marka).{N}")
        print(f"  Çakışan {len(cakisma)} alana dokunulmadı.")
        return 0
    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(main())
